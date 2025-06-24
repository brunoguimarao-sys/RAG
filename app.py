import os
import sqlite3
import time
import logging
from pathlib import Path
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from werkzeug.utils import secure_filename
import threading
import fitz # PyMuPDF
from docx2python import docx2python
from PIL import Image
import pytesseract
import cv2
import numpy as np
import requests
from sentence_transformers import SentenceTransformer
from langchain.text_splitter import RecursiveCharacterTextSplitter
import chromadb
from concurrent.futures import ProcessPoolExecutor, as_completed
import sys
import subprocess
import atexit
import torch
import openpyxl

# ==============================================================================
# --- SEÇÃO: CONFIGURAÇÃO DE LOGGING ---
# ==============================================================================
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s - %(message)s')
# File handler para escrever logs em um arquivo
file_handler = logging.FileHandler('rag_app.log', mode='w', encoding='utf-8')
file_handler.setLevel(logging.INFO)
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)
# Stream handler para exibir logs no console
stream_handler = logging.StreamHandler()
stream_handler.setLevel(logging.INFO)
stream_handler.setFormatter(formatter)
logger.addHandler(stream_handler)

# ==============================================================================
# --- SEÇÃO: CAMINHOS E PARÂMETROS GLOBAIS ---
# ==============================================================================
UPLOAD_FOLDER = 'uploads'
DOCUMENTS_DIR = os.getenv("DOCUMENTS_DIR", "C:\\Documentos")
CHROMA_DB_PATH = os.getenv("CHROMA_DB_PATH", "chroma_db")
CHROMA_COLLECTION_NAME = "document_chunks"
DB_PATH = "audit.db" # Banco de dados SQLite para auditoria
TESSERACT_CMD_PATH = os.getenv("TESSERACT_CMD_PATH", r"C:\Users\bruno.guimaraes\AppData\Local\Programs\Tesseract-OCR\tesseract.exe")

# ==============================================================================
# --- SEÇÃO: PARÂMETROS DO SERVIDOR LLAMA ---
# ==============================================================================
LLAMA_MODEL_PATH = os.getenv("LLAMA_MODEL_PATH", r"C:\LLM\amoral-gemma3-4B-v2-qat.Q4_K_S.gguf")
LLAMA_SERVER_URL = os.getenv("LLAMA_SERVER_URL", "http://localhost:8080/completion")
LLAMA_HOST = "127.0.0.1"
LLAMA_PORT = 8080
LLAMA_NGL = -1 # Número de camadas para descarregar na GPU (-1 para GPU se disponível, 0 para CPU)
LLAMA_THREADS = 10
LLAMA_THREADS_BATCH = 10
LLAMA_BATCH_SIZE = 512
LLAMA_CONTEXT_SIZE = 4096
LLAMA_TEMP = 0.7
LLAMA_TOP_K = 40
LLAMA_TOP_P = 0.95
LLAMA_REPEAT_PENALTY = 1.1
llama_server_process = None # Variável global para o processo do servidor Llama

# ==============================================================================
# --- SEÇÃO: PARÂMETROS DE EMBEDDING E INDEXAÇÃO ---
# ==============================================================================
EMBEDDING_MODEL = "all-MiniLM-L6-v2" # Modelo de embedding da Sentence Transformers
CHUNK_SIZE = 500 # Tamanho dos chunks de texto
CHUNK_OVERLAP = 50 # Sobreposição entre chunks
TOP_K = 3 # Número de chunks mais relevantes a serem recuperados
MAX_WORKERS = os.cpu_count() or 4 # Número de workers para processamento paralelo
SUPPORTED_EXTENSIONS = {'.pdf', '.docx', '.txt', '.jpg', '.jpeg', '.png', '.xlsx'} # Adicionado '.xlsx'

# ==============================================================================
# --- SEÇÃO: INICIALIZAÇÃO DE APLICAÇÃO E MODELOS ---
# ==============================================================================
app = Flask(__name__, static_folder="static")
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
CORS(app) # Habilita CORS para a aplicação Flask

# Define o dispositivo para o modelo de embedding (GPU se disponível, senão CPU)
device = 'cuda' if torch.cuda.is_available() else 'cpu'
logger.info(f"Usando dispositivo para embedding: {device.upper()}")

logger.info("Carregando o modelo de embedding...")
embedding_model = SentenceTransformer(EMBEDDING_MODEL, device=device)
logger.info("Modelo de embedding carregado com sucesso.")

# Configura o caminho do Tesseract OCR, se existir
if os.path.exists(TESSERACT_CMD_PATH):
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD_PATH
else:
    logger.error(f"Caminho do Tesseract OCR não encontrado: {TESSERACT_CMD_PATH}. A extração de texto de imagens pode falhar.")

# ==============================================================================
# --- SEÇÃO: GERENCIAMENTO DO SERVIDOR LLAMA ---
# ==============================================================================
def is_server_running():
    """Verifica se o servidor Llama está em execução."""
    try:
        requests.get(f"http://{LLAMA_HOST}:{LLAMA_PORT}/", timeout=2)
        return True
    except requests.exceptions.ConnectionError:
        return False

def start_llama_server():
    """Inicia o servidor Llama em um processo separado."""
    global llama_server_process
    if is_server_running():
        logger.info("Servidor Llama já está em execução.")
        return

    if not os.path.exists(LLAMA_MODEL_PATH):
        logger.critical(f"CRÍTICO: Modelo do LLM não encontrado em: '{LLAMA_MODEL_PATH}'. O servidor não pode ser iniciado.")
        sys.exit(1) # Encerra a aplicação se o modelo não for encontrado

    command = [
        "llama-server", "-m", LLAMA_MODEL_PATH, "-c", str(LLAMA_CONTEXT_SIZE),
        "-t", str(LLAMA_THREADS), "-tb", str(LLAMA_THREADS_BATCH), "-b", str(LLAMA_BATCH_SIZE),
        "--host", LLAMA_HOST, "--port", str(LLAMA_PORT), "-ngl", str(LLAMA_NGL),
        "--temp", str(LLAMA_TEMP), "--top-k", str(LLAMA_TOP_K),
        "--top-p", str(LLAMA_TOP_P), "--repeat-penalty", str(LLAMA_REPEAT_PENALTY)
    ]
    logger.info(f"Iniciando o servidor Llama com o comando: {' '.join(command)}")
    # Inicia o processo do servidor Llama
    llama_server_process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, encoding='utf-8', errors='replace')
    
    logger.info("Aguardando o servidor Llama inicializar (até 30 segundos)...")
    time.sleep(10) # Dá um tempo inicial para o servidor começar
    for _ in range(20): # Tenta verificar por mais 20 segundos
        if is_server_running():
            logger.info("Servidor Llama iniciado e respondendo com sucesso!")
            return
        if llama_server_process.poll() is not None: # Verifica se o processo terminou
            break 
        time.sleep(1)

    logger.error("O servidor Llama não iniciou corretamente ou encerrou inesperadamente.")
    if llama_server_process:
        llama_server_process.terminate()
        try:
            output, _ = llama_server_process.communicate(timeout=5) # Tenta obter a saída do processo
            logger.error("SAÍDA DO PROCESSO DO SERVIDOR LLAMA (ERRO):\n" + "="*50 + f"\n{output}\n" + "="*50)
        except subprocess.TimeoutExpired:
            logger.error("Não foi possível obter a saída do processo do servidor Llama (timeout).")
    sys.exit(1) # Encerra a aplicação se o servidor Llama não iniciar

def shutdown_llama_server():
    """Encerra o processo do servidor Llama."""
    global llama_server_process
    if llama_server_process and llama_server_process.poll() is None: # Verifica se o processo existe e está em execução
        logger.info("Encerrando o servidor Llama...")
        llama_server_process.terminate()
        try:
            llama_server_process.wait(timeout=10) # Espera o processo terminar
            logger.info("Servidor Llama encerrado com sucesso.")
        except subprocess.TimeoutExpired:
            logger.warning("Servidor Llama não encerrou a tempo, forçando o encerramento (kill).")
            llama_server_process.kill() # Força o encerramento se o terminate não funcionar

atexit.register(shutdown_llama_server) # Garante que o servidor Llama seja encerrado ao sair da aplicação

# ==============================================================================
# --- SEÇÃO: BANCO DE DADOS (CHROMA DB E SQLITE) E AUDITORIA ---
# ==============================================================================
client = chromadb.PersistentClient(path=CHROMA_DB_PATH) # Cliente persistente para o ChromaDB
collection = None # Variável global para a coleção do ChromaDB

def get_or_create_collection():
    """Obtém ou cria a coleção no ChromaDB com uma função de embedding customizada."""
    global collection
    if collection is None:
        class CustomEmbeddingFunction(chromadb.api.types.EmbeddingFunction):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
            def __call__(self, texts: list[str]) -> list[list[float]]:
                # Gera embeddings em lotes para eficiência
                return embedding_model.encode(texts, batch_size=128, show_progress_bar=False).tolist()

        collection = client.get_or_create_collection(
            name=CHROMA_COLLECTION_NAME,
            embedding_function=CustomEmbeddingFunction() # Usa a função de embedding customizada
        )
        logger.info(f"Coleção '{CHROMA_COLLECTION_NAME}' carregada/criada no ChromaDB.")
    return collection

def init_db():
    """Inicializa o banco de dados SQLite para auditoria."""
    try:
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS auditoria (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    pergunta TEXT,
                    resposta TEXT,
                    contexto TEXT,
                    arquivos_utilizados TEXT,
                    data TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
        logger.info("Banco de dados de auditoria inicializado com sucesso.")
    except sqlite3.Error as e:
        logger.error(f"Erro ao inicializar o banco de dados de auditoria: {e}")

def registrar_auditoria(pergunta: str, resposta: str, contexto: str, arquivos: list[str]):
    """Registra uma entrada de auditoria no banco de dados SQLite."""
    try:
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute(
                "INSERT INTO auditoria (pergunta, resposta, contexto, arquivos_utilizados) VALUES (?,?,?,?)",
                (pergunta, resposta, contexto, '; '.join(arquivos)) # Concatena a lista de arquivos em uma string
            )
            conn.commit()
        logger.info(f"Auditoria registrada para a pergunta: '{pergunta[:50]}...'")
    except sqlite3.Error as e:
        logger.error(f"Erro ao registrar auditoria: {e}")

# ==============================================================================
# --- SEÇÃO: PROCESSAMENTO E EXTRAÇÃO DE TEXTO DE ARQUIVOS ---
# ==============================================================================
def encontrar_arquivos_recursivamente(diretorio: str) -> list[Path]:
    """Encontra todos os arquivos com extensões suportadas em um diretório recursivamente."""
    arquivos_encontrados = []
    diretorio_path = Path(diretorio)
    if not diretorio_path.exists() or not diretorio_path.is_dir():
        logger.error(f"Diretório de documentos não existe ou não é um diretório: {diretorio}")
        return arquivos_encontrados
    
    logger.info(f"Buscando arquivos recursivamente em: {diretorio}")
    for extensao in SUPPORTED_EXTENSIONS:
        pattern = f"**/*{extensao}" # Padrão para encontrar arquivos com a extensão
        arquivos = list(diretorio_path.rglob(pattern))
        arquivos_encontrados.extend(arquivos)
        if arquivos:
            logger.info(f"Encontrados {len(arquivos)} arquivos com extensão {extensao}")
            
    arquivos_encontrados = sorted(list(set(arquivos_encontrados))) # Remove duplicatas e ordena
    logger.info(f"Total de arquivos únicos encontrados: {len(arquivos_encontrados)}")
    return arquivos_encontrados

# Funções de pré-processamento de imagem para OCR
def get_grayscale(image: np.ndarray) -> np.ndarray:
    """Converte uma imagem para escala de cinza."""
    return cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

def deskew(image: np.ndarray) -> np.ndarray:
    """Corrige a inclinação (skew) de uma imagem."""
    if len(image.shape) == 3: image = get_grayscale(image) # Converte para cinza se necessário
    coords = np.column_stack(np.where(image > 0)) # Encontra coordenadas de pixels não pretos
    if len(coords) < 2: return image # Retorna a imagem original se não houver pixels suficientes
    
    angle = cv2.minAreaRect(coords)[-1] # Calcula o ângulo de inclinação
    if angle < -45: angle = -(90 + angle)
    else: angle = -angle
    
    (h, w) = image.shape[:2]
    center = (w // 2, h // 2)
    M = cv2.getRotationMatrix2D(center, angle, 1.0) # Cria a matriz de rotação
    # Aplica a rotação para corrigir a inclinação
    return cv2.warpAffine(image, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)

def thresholding(image: np.ndarray) -> np.ndarray:
    """Aplica binarização (thresholding) à imagem."""
    if len(image.shape) == 3: image = get_grayscale(image) # Converte para cinza se necessário
    # Usa o método de Otsu para binarização automática
    return cv2.threshold(image, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]

def extrair_texto(path: Path) -> str:
    """Extrai texto de diferentes tipos de arquivo (PDF, DOCX, TXT, Imagens, XLSX)."""
    texto_completo = ""
    try:
        suffix = path.suffix.lower() # Obtém a extensão do arquivo em minúsculas
        
        if suffix == ".pdf":
            with fitz.open(path) as doc:
                for page_num, page in enumerate(doc):
                    text_native = page.get_text().strip()
                    if text_native: # Tenta extrair texto nativo primeiro
                        texto_completo += text_native + "\n"
                    else: # Se não houver texto nativo, tenta OCR
                        logger.info(f"Página {page_num+1} do PDF '{path.name}' não contém texto nativo, tentando OCR.")
                        pix = page.get_pixmap(dpi=300) # Renderiza a página como imagem
                        img_np = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.h, pix.w, pix.n)
                        if pix.n == 4: img_np = cv2.cvtColor(img_np, cv2.COLOR_RGBA2BGR) # Converte RGBA para BGR
                        
                        # Pré-processamento da imagem para OCR
                        img_processed = get_grayscale(img_np)
                        img_processed = thresholding(img_processed)
                        img_processed = deskew(img_processed)
                        
                        texto_ocr = pytesseract.image_to_string(Image.fromarray(img_processed), lang='por+eng') # OCR com Português e Inglês
                        if texto_ocr.strip():
                            logger.info(f"Texto extraído via OCR da página {page_num+1} do PDF '{path.name}'.")
                            texto_completo += texto_ocr + "\n"
                        else:
                            logger.warning(f"Nenhum texto extraído (OCR) da página {page_num+1} do PDF '{path.name}'.")
        elif suffix == ".docx":
            with docx2python(str(path)) as docx_content:
                texto_completo = docx_content.text
        elif suffix == ".txt":
            # Tenta diferentes encodings comuns para arquivos de texto
            for encoding in ['utf-8', 'latin-1', 'cp1252']:
                try:
                    texto_completo = path.read_text(encoding=encoding)
                    logger.info(f"Texto extraído de '{path.name}' com encoding '{encoding}'.")
                    break 
                except UnicodeDecodeError:
                    if encoding == 'cp1252': # Se falhar no último encoding, loga o erro
                        logger.warning(f"Não foi possível decodificar '{path.name}' com encodings testados.")
                    continue
        elif suffix in [".jpg", ".jpeg", ".png"]:
            img = cv2.imread(str(path))
            if img is None: 
                logger.error(f"Não foi possível carregar a imagem: {path.name}")
                raise ValueError("Não foi possível carregar a imagem")
            
            # Pré-processamento da imagem para OCR
            img_processed = get_grayscale(img)
            img_processed = thresholding(img_processed)
            img_processed = deskew(img_processed)
            
            texto_completo = pytesseract.image_to_string(Image.fromarray(img_processed), lang='por+eng') # OCR
            if not texto_completo.strip():
                logger.warning(f"Nenhum texto extraído (OCR) da imagem '{path.name}'.")
        elif suffix == ".xlsx": # NOVO: Suporte a XLSX
            logger.info(f"Extraindo texto do arquivo Excel: '{path.name}'")
            try:
                workbook = openpyxl.load_workbook(path, data_only=True) # data_only=True para obter valores, não fórmulas
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    texto_completo += f"\n--- Planilha: {sheet_name} ---\n"
                    for row in sheet.iter_rows():
                        row_values = []
                        for cell in row:
                            if cell.value is not None:
                                row_values.append(str(cell.value).strip())
                        if row_values:
                            texto_completo += " ".join(row_values) + "\n"
                logger.info(f"Texto extraído de '{path.name}' (XLSX).")
            except Exception as e:
                logger.error(f"Erro ao ler arquivo XLSX '{path.name}': {e}", exc_info=True)
                return "" # Retorna vazio em caso de erro na leitura do XLSX

    except Exception as e:
        logger.error(f"Falha ao extrair texto de {path.name}: {e}", exc_info=True) # Loga a exceção com traceback
        return "" # Retorna string vazia em caso de erro
    
    return texto_completo.strip()

# ==============================================================================
# --- SEÇÃO: INDEXAÇÃO E GERAÇÃO DE CHUNKS ---
# ==============================================================================
def quebrar_em_chunks(texto: str, arquivo_path: str) -> tuple[list[str], list[dict], list[str]]:
    """Quebra o texto em chunks menores e gera metadados e IDs para cada chunk."""
    if not texto.strip(): return [], [], [] # Retorna listas vazias se o texto estiver vazio
    
    # Usa RecursiveCharacterTextSplitter para quebrar o texto
    splitter = RecursiveCharacterTextSplitter(chunk_size=CHUNK_SIZE, chunk_overlap=CHUNK_OVERLAP, length_function=len)
    chunks_text = splitter.split_text(texto)
    
    metadados_list, ids_list = [], []
    # Normaliza o caminho do arquivo para criar IDs únicos e consistentes
    normalized_file_path = str(Path(arquivo_path)).encode('utf-8').hex() 
    
    for i, chunk_content in enumerate(chunks_text):
        unique_chunk_id = f"{normalized_file_path}_chunk_{i}" # Cria um ID único para o chunk
        ids_list.append(unique_chunk_id)
        metadados_list.append({
            'arquivo_path': str(arquivo_path), 
            'arquivo_nome': Path(arquivo_path).name, 
            'chunk_id': i
        })
    return chunks_text, metadados_list, ids_list

def processar_arquivo_para_indexacao(arquivo_path: Path) -> tuple[list[str] | None, list[dict] | None, list[str] | None, bool]:
    """Processa um único arquivo: extrai texto, quebra em chunks e prepara para indexação."""
    try:
        logger.info(f"Processando arquivo para indexação: {arquivo_path.name}")
        texto = extrair_texto(arquivo_path)
        if not texto:
            logger.warning(f"Nenhum texto válido extraído de: {arquivo_path.name}. Arquivo não será indexado.")
            return None, None, None, False
            
        chunks, metadatas, ids = quebrar_em_chunks(texto, str(arquivo_path))
        if not chunks:
            logger.warning(f"Texto extraído de '{arquivo_path.name}' não gerou chunks válidos. Arquivo não será indexado.")
            return None, None, None, False
            
        logger.info(f"[{len(chunks)}] chunks gerados com sucesso de {arquivo_path.name}")
        return chunks, metadatas, ids, True
    except Exception as e:
        logger.error(f"Erro crítico ao processar {arquivo_path.name} para indexação: {e}", exc_info=True)
        return None, None, None, False

def indexar_documentos(diretorio_documentos: str = DOCUMENTS_DIR) -> bool:
    """Indexa todos os documentos encontrados no diretório especificado."""
    logger.info(f"Iniciando processo de indexação completa dos documentos em: {diretorio_documentos}")
    arquivos_para_indexar = encontrar_arquivos_recursivamente(diretorio_documentos)
    
    if not arquivos_para_indexar:
        logger.error("Nenhum arquivo encontrado para indexação. Verifique o diretório e as extensões suportadas.")
        return False

    logger.info("Recriando coleção no ChromaDB para garantir uma reindexação completa...")
    try:
        client.delete_collection(name=CHROMA_COLLECTION_NAME) # Deleta a coleção existente para evitar duplicatas
        logger.info(f"Coleção antiga '{CHROMA_COLLECTION_NAME}' deletada com sucesso.")
    except Exception as e: # chromadb.errors.CollectionNotFoundError ou similar
        logger.warning(f"Não foi possível deletar a coleção '{CHROMA_COLLECTION_NAME}' (provavelmente não existia): {e}")
    
    chroma_collection = get_or_create_collection() # Cria uma nova coleção (ou obtém se já existir, embora deletada acima)
    
    all_chunks_list, all_metadatas_list, all_ids_list = [], [], []
    arquivos_processados_sucesso, arquivos_processados_erro = 0, 0

    # Usa ProcessPoolExecutor para processar arquivos em paralelo
    with ProcessPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_file = {executor.submit(processar_arquivo_para_indexacao, arq_path): arq_path for arq_path in arquivos_para_indexar}
        for future in as_completed(future_to_file): # Processa os resultados à medida que ficam prontos
            try:
                chunks, metadatas, ids, sucesso = future.result()
                if sucesso and chunks: # Verifica se o processamento foi bem-sucedido e gerou chunks
                    all_chunks_list.extend(chunks)
                    all_metadatas_list.extend(metadatas)
                    all_ids_list.extend(ids)
                    arquivos_processados_sucesso += 1
                else:
                    arquivos_processados_erro += 1
            except Exception as exc:
                arquivo_orig = future_to_file[future]
                logger.error(f"Falha no processamento do arquivo {arquivo_orig.name} durante a indexação em paralelo: {exc}", exc_info=True)
                arquivos_processados_erro += 1
                
    logger.info(f"Processamento de arquivos para indexação concluído. Sucesso: {arquivos_processados_sucesso}, Erros: {arquivos_processados_erro}.")

    if not all_chunks_list:
        logger.warning("Nenhum chunk válido foi gerado para indexação após o processamento de todos os arquivos.")
        return False

    logger.info(f"Adicionando {len(all_chunks_list)} chunks ao ChromaDB em lotes...")
    try:
        # Adiciona os chunks ao ChromaDB em lotes para otimizar o desempenho
        batch_size_chroma = 500 # Tamanho do lote para adicionar ao ChromaDB
        for i in range(0, len(all_chunks_list), batch_size_chroma):
            chroma_collection.add(
                documents=all_chunks_list[i:i+batch_size_chroma],
                metadatas=all_metadatas_list[i:i+batch_size_chroma],
                ids=all_ids_list[i:i+batch_size_chroma]
            )
            logger.info(f"Lote {i//batch_size_chroma + 1} de { (len(all_chunks_list) + batch_size_chroma -1) // batch_size_chroma } adicionado ao ChromaDB.")
        logger.info("Indexação de todos os documentos concluída com sucesso!")
        return True
    except Exception as e:
        logger.error(f"Falha crítica ao adicionar documentos ao ChromaDB: {e}", exc_info=True)
        return False

def indexar_arquivo_unico(caminho_arquivo: str):
    """Processa e indexa um único arquivo no ChromaDB. Usado para uploads."""
    try:
        logger.info(f"[INDEXAÇÃO DE ARQUIVO ÚNICO EM BACKGROUND] Iniciando para: {caminho_arquivo}")
        arquivo_path = Path(caminho_arquivo)
        
        if not arquivo_path.exists():
            logger.error(f"[INDEXAÇÃO EM BACKGROUND] Arquivo não encontrado: {caminho_arquivo}")
            return

        chunks, metadatas, ids, sucesso = processar_arquivo_para_indexacao(arquivo_path)

        if not sucesso or not chunks:
            logger.error(f"[INDEXAÇÃO EM BACKGROUND] Falha ao processar e gerar chunks para {arquivo_path.name}. O arquivo não será indexado.")
            return

        chroma_collection = get_or_create_collection() # Garante que a coleção esteja disponível
        # Adiciona os chunks do arquivo único à coleção existente
        chroma_collection.add(
            documents=chunks,
            metadatas=metadatas,
            ids=ids
        )
        logger.info(f"[INDEXAÇÃO EM BACKGROUND] Sucesso! Adicionados {len(chunks)} chunks de {arquivo_path.name} ao ChromaDB.")

    except Exception as e:
        logger.error(f"[INDEXAÇÃO EM BACKGROUND] Erro crítico ao indexar o arquivo '{caminho_arquivo}': {e}", exc_info=True)

# ==============================================================================
# --- SEÇÃO: RECUPERAÇÃO DE CONTEXTO E GERAÇÃO DE RESPOSTA (RAG) ---
# ==============================================================================
def recuperar_contexto(query: str, top_k_chunks: int = TOP_K) -> tuple[str, list[str]]:
    """Recupera chunks de contexto relevantes do ChromaDB com base na query."""
    try:
        chroma_collection = get_or_create_collection()
        if chroma_collection.count() == 0:
            logger.warning("O índice (ChromaDB) está vazio. Execute a reindexação primeiro ou faça upload de arquivos.")
            return "", [] # Retorna contexto vazio e lista de arquivos vazia
            
        # Realiza a query na coleção do ChromaDB
        results = chroma_collection.query(
            query_texts=[query], 
            n_results=top_k_chunks, 
            include=['documents', 'metadatas'] # Inclui os documentos (chunks) e metadados
        )
        
        chunks_com_fonte, arquivos_utilizados_set = [], set()
        if results and results['documents'] and results['documents'][0]:
            document_list, metadata_list = results['documents'][0], results['metadatas'][0]
            for doc_content, meta_info in zip(document_list, metadata_list):
                arquivo_nome_orig = meta_info.get('arquivo_nome', 'Desconhecido')
                arquivos_utilizados_set.add(arquivo_nome_orig)
                chunks_com_fonte.append(f"[Fonte: {arquivo_nome_orig}]\n{doc_content}") # Adiciona a fonte ao chunk
        
        contexto_final = "\n---\n".join(chunks_com_fonte) # Une os chunks com um separador
        return contexto_final, list(arquivos_utilizados_set)
        
    except Exception as e:
        logger.error(f"Falha ao recuperar contexto do ChromaDB: {e}", exc_info=True)
        return f"Erro ao recuperar contexto do ChromaDB: {e}", [] # Retorna mensagem de erro e lista vazia

def gerar_resposta(prompt_completo: str) -> str:
    """Envia o prompt completo para o servidor Llama e obtém a resposta."""
    if not is_server_running():
        logger.error("Servidor Llama não está em execução. Não é possível gerar resposta.")
        return "Erro: O servidor de IA (Llama) não está disponível no momento. Por favor, tente mais tarde."

    try:
        payload = {
            "prompt": prompt_completo, 
            "n_predict": 1024, # Aumentado para respostas potencialmente mais longas
            "temperature": LLAMA_TEMP, 
            "top_k": LLAMA_TOP_K,
            "top_p": LLAMA_TOP_P,
            "repeat_penalty": LLAMA_REPEAT_PENALTY,
            "stop": ["\n\n", "Pergunta:", "Contexto:", "Usuário:"] # Palavras/frases para parar a geração
        }
        response = requests.post(LLAMA_SERVER_URL, json=payload, timeout=120) # Timeout aumentado para 2 minutos
        response.raise_for_status() # Levanta uma exceção para códigos de erro HTTP (4xx ou 5xx)
        
        resposta_json = response.json()
        # Verifica se a resposta contém 'content' e se não está vazia
        if "content" in resposta_json and resposta_json["content"].strip():
            return resposta_json["content"].strip()
        else:
            logger.warning(f"Resposta do Llama não continha 'content' ou estava vazia. Payload: {payload}, Resposta: {resposta_json}")
            return "O modelo de IA não forneceu uma resposta válida para esta pergunta."
            
    except requests.exceptions.Timeout:
        logger.error(f"Timeout na comunicação com o servidor Llama em {LLAMA_SERVER_URL}.")
        return f"Erro: Timeout ao tentar conectar ao servidor de IA. A solicitação demorou muito para responder."
    except requests.exceptions.RequestException as e:
        logger.error(f"Falha na comunicação com o servidor Llama: {e}")
        return f"Erro: Não foi possível conectar ao servidor de IA em {LLAMA_SERVER_URL}. Detalhes: {e}"
    except Exception as e:
        logger.error(f"Falha inesperada ao gerar resposta: {e}", exc_info=True)
        return f"Erro inesperado ao gerar a resposta: {e}"

# ==============================================================================
# --- SEÇÃO: ENDPOINTS DA API FLASK ---
# ==============================================================================
@app.route("/")
def frontend():
    """Serve a página principal do frontend (index.html)."""
    return send_from_directory("static", "index.html")

@app.route("/perguntar", methods=['POST'])
def perguntar_endpoint():
    """Endpoint para receber perguntas, processá-las e retornar respostas."""
    try:
        # Lida com 'multipart/form-data' para permitir upload de arquivos junto com a pergunta
        query = request.form.get("pergunta", "").strip()
        uploaded_file = request.files.get("arquivo") # 'arquivo' é o nome do campo no form-data
        incluir_contexto_na_resposta = request.form.get("incluir_contexto", 'false').lower() == 'true'
        
        if not query:
            return jsonify({"erro": "O campo 'pergunta' não pode estar vazio."}), 400

        logger.info(f"Recebida pergunta: '{query}' (Incluir contexto: {incluir_contexto_na_resposta})")
        
        contexto_do_upload = ""
        arquivos_utilizados_no_upload = []

        # Processa o arquivo enviado, se houver
        if uploaded_file and uploaded_file.filename:
            filename = secure_filename(uploaded_file.filename) # Garante um nome de arquivo seguro
            # Cria a pasta de uploads se não existir
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            uploaded_file.save(save_path)
            logger.info(f"Arquivo '{filename}' recebido e salvo em '{save_path}'.")
            
            # 1. Extrai texto do arquivo para uso imediato no contexto desta pergunta
            logger.info(f"Extraindo texto do arquivo '{filename}' para contexto imediato.")
            contexto_do_upload = extrair_texto(Path(save_path))
            if contexto_do_upload:
                arquivos_utilizados_no_upload.append(filename)
                logger.info(f"Texto extraído de '{filename}' para contexto: {len(contexto_do_upload)} caracteres.")
            else:
                logger.warning(f"Nenhum texto extraído do arquivo '{filename}' para contexto imediato.")

            # 2. Inicia a indexação do arquivo em uma thread separada (não bloqueia a resposta)
            logger.info(f"Iniciando a indexação de '{filename}' em segundo plano.")
            index_thread = threading.Thread(target=indexar_arquivo_unico, args=(save_path,))
            index_thread.start() # A indexação ocorrerá em background

        # Recupera contexto da base de dados (ChromaDB) existente
        logger.info(f"Recuperando contexto da base de dados para a query: '{query}'")
        contexto_da_base, arquivos_utilizados_na_base = recuperar_contexto(query)
        if contexto_da_base:
            logger.info(f"Contexto recuperado da base: {len(contexto_da_base)} caracteres, de {len(arquivos_utilizados_na_base)} arquivos.")
        else:
            logger.info("Nenhum contexto relevante encontrado na base de dados para esta query.")

        # Combina os contextos (upload e base)
        contextos_combinados = []
        if contexto_do_upload:
            contextos_combinados.append(f"--- CONTEXTO DO ARQUIVO '{uploaded_file.filename if uploaded_file else 'N/A'}' (Recém Enviado) ---\n{contexto_do_upload}")
        if contexto_da_base:
            contextos_combinados.append(f"--- CONTEXTO DA BASE DE DADOS EXISTENTE ---\n{contexto_da_base}")
        
        contexto_final_para_prompt = "\n\n".join(contextos_combinados)
        todos_arquivos_utilizados = list(set(arquivos_utilizados_no_upload + arquivos_utilizados_na_base))
        
        # Monta o prompt para o modelo Llama
        prompt_final = f"""Você é um assistente de IA especializado em responder perguntas com base em documentos e informações fornecidas.

Contexto dos documentos (se disponível):
{contexto_final_para_prompt if contexto_final_para_prompt else "Nenhum contexto específico foi encontrado para esta pergunta."}

Pergunta do usuário: {query}

Instruções para a resposta:
- Responda de forma clara, concisa e objetiva em Português do Brasil.
- Utilize APENAS as informações do contexto fornecido. Se o contexto do arquivo recém-enviado e o da base de dados estiverem disponíveis e forem relevantes, você pode usar ambos, mas dê preferência ou destaque a informação do arquivo recém-adicionado se houver conflito ou sobreposição.
- Se a informação necessária para responder à pergunta não estiver no contexto, responda educadamente: "A informação não foi encontrada nos documentos consultados." ou "Com base nas informações disponíveis, não consigo responder a essa pergunta."
- NÃO invente informações ou use conhecimento externo.
- Ao final da sua resposta principal, se você utilizou informações de algum arquivo, liste os nomes dos arquivos fonte que foram efetivamente usados para formular a resposta, no formato: [Fonte(s) utilizada(s): nome_do_arquivo1.pdf; nome_do_arquivo2.txt]. Se nenhum arquivo específico do contexto foi usado, omita esta parte.

Resposta:"""

        logger.info(f"Enviando prompt para o Llama. Tamanho do prompt: {len(prompt_final)} caracteres.")
        resposta_gerada = gerar_resposta(prompt_final)
        
        # Registra a auditoria da pergunta e resposta
        registrar_auditoria(query, resposta_gerada, contexto_final_para_prompt, todos_arquivos_utilizados)

        return jsonify({
            "pergunta": query,
            "resposta": resposta_gerada,
            "arquivos_consultados": todos_arquivos_utilizados,
            "contexto_utilizado": contexto_final_para_prompt if incluir_contexto_na_resposta else None # Retorna o contexto apenas se solicitado
        })
        
    except Exception as e:
        logger.error(f"Falha crítica na API /perguntar: {e}", exc_info=True)
        return jsonify({"erro": f"Erro interno no servidor ao processar a pergunta: {e}"}), 500

@app.route("/status", methods=['GET'])
def status_endpoint():
    """Endpoint para verificar o status da aplicação e do servidor Llama."""
    try:
        chroma_collection_local = get_or_create_collection()
        total_chunks_indexados = chroma_collection_local.count() if chroma_collection_local else 0
        
        return jsonify({
            "servidor_flask_online": True,
            "servidor_llm_online": is_server_running(),
            "diretorio_documentos_configurado": DOCUMENTS_DIR,
            "diretorio_documentos_existe": os.path.exists(DOCUMENTS_DIR) and os.path.isdir(DOCUMENTS_DIR),
            "total_chunks_indexados_chromadb": total_chunks_indexados,
            "colecao_chromadb_vazia": total_chunks_indexados == 0,
            "modelo_embedding_carregado": embedding_model is not None,
            "tesseract_configurado": os.path.exists(pytesseract.pytesseract.tesseract_cmd) if pytesseract.pytesseract.tesseract_cmd else False
        })
    except Exception as e:
        logger.error(f"Erro ao verificar status da aplicação: {e}", exc_info=True)
        return jsonify({"erro": f"Erro ao verificar status: {e}"}), 500

@app.route("/reindexar", methods=['POST'])
def reindexar_endpoint():
    """Endpoint para acionar a reindexação completa dos documentos."""
    logger.info("Requisição de reindexação completa recebida via API.")
    try:
        # Adicionar verificação de segurança aqui se necessário (ex: token de admin)
        sucesso_reindexacao = indexar_documentos() # Chama a função principal de indexação
        if sucesso_reindexacao:
            return jsonify({"mensagem": "Reindexação completa dos documentos concluída com sucesso!"})
        else:
            return jsonify({"erro": "Falha no processo de reindexação completa. Verifique os logs para mais detalhes."}), 500
    except Exception as e:
        logger.error(f"Erro crítico durante a chamada da API /reindexar: {e}", exc_info=True)
        return jsonify({"erro": f"Erro crítico durante a reindexação via API: {e}"}), 500

# ==============================================================================
# --- SEÇÃO: BLOCO PRINCIPAL DE EXECUÇÃO (MAIN) ---
# ==============================================================================
if __name__ == "__main__":
    # Garante que a pasta de uploads exista ao iniciar a aplicação
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    logger.info(f"Pasta de uploads '{UPLOAD_FOLDER}' verificada/criada.")

    # Verifica se o script foi chamado com o argumento "indexar"
    if len(sys.argv) > 1 and sys.argv[1].lower() == "indexar":
        logger.info("=== MODO DE INDEXAÇÃO INICIADO VIA LINHA DE COMANDO ===")
        init_db() # Inicializa o DB de auditoria
        # Não precisa de get_or_create_collection aqui, pois indexar_documentos faz isso.
        sucesso_idx = indexar_documentos() # Chama a função de indexação
        if sucesso_idx:
            logger.info("\n=== INDEXAÇÃO CONCLUÍDA COM SUCESSO (VIA LINHA DE COMANDO) ===")
            sys.exit(0) # Sai com sucesso
        else:
            logger.error("\n=== INDEXAÇÃO FALHOU (VIA LINHA DE COMANDO) ===")
            sys.exit(1) # Sai com erro
    else:
        # Modo normal de execução (servidor Flask)
        init_db() # Inicializa o DB de auditoria
        get_or_create_collection() # Garante que a coleção ChromaDB esteja pronta
        
        # Tenta iniciar o servidor Llama. Se falhar, start_llama_server já trata o sys.exit.
        start_llama_server() 
        
        logger.info("=== SERVIDOR RAG INICIADO EM MODO DE APLICAÇÃO ===")
        logger.info(f"Servidor Flask estará disponível em http://{'0.0.0.0'}:5000")
        logger.info(f"Para reindexar todos os documentos, execute: python {Path(__file__).name} indexar")
        # Executa a aplicação Flask (debug=False para produção/uso normal)
        app.run(host='0.0.0.0', port=5000, debug=False)